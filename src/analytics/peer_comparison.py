import os
import sqlite3

import pandas as pd

DB_PATH = "db/nifty100.db"

conn = sqlite3.connect(DB_PATH)

financial_data = pd.read_sql("SELECT * FROM financial_data", conn)

financial_ratios = pd.read_sql("SELECT * FROM financial_ratios", conn)

peer_percentiles = pd.read_sql("SELECT * FROM peer_percentiles", conn)

print("Financial Data:", financial_data.shape)
print("Financial Ratios:", financial_ratios.shape)
print("Peer Percentiles:", peer_percentiles.shape)
merged_df = financial_data.merge(
    financial_ratios, on=["company_id", "year"], how="inner"
)

merged_df = merged_df.merge(peer_percentiles, on=["company_id", "year"], how="left")

print("Merged Shape:", merged_df.shape)

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

OUTPUT_FILE = os.path.join(OUTPUT_DIR, "peer_comparison.xlsx")

writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")

peer_groups = sorted(peer_percentiles["peer_group_name"].dropna().unique())

print("Peer Groups:", len(peer_groups))
print(peer_groups)

for group in peer_groups:
    group_df = merged_df[merged_df["peer_group_name"] == group].copy()

    sheet_name = group[:31]  # Excel sheet name limit

    group_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Written sheet: {sheet_name} ({len(group_df)} rows)")

writer.close()

print(f"\nExcel report saved to: {OUTPUT_FILE}")


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(OUTPUT_FILE)

green_fill = PatternFill(fill_type="solid", fgColor="90EE90")
yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF99")
red_fill = PatternFill(fill_type="solid", fgColor="FF9999")

for sheet in wb.sheetnames:
    ws = wb[sheet]

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value >= 80:
                    cell.fill = green_fill
                elif cell.value >= 50:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill

wb.save(OUTPUT_FILE)

print("Excel formatting completed successfully!")
